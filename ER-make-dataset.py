import networkx as nx
import pandas as pd

# ノード数 : n
# エッジの張り替える確率 : p

degrees = []
eigenvectors = []
pageranks = []
betweenesses = []

# 次数集中度
def degree_concentrate(G):
    # 次数中心性
    gd_degree_centrality = pd.Series(nx.degree_centrality(G))
    sum = 0
    for i in range(0, len(gd_degree_centrality)):
        sum += gd_degree_centrality[i]
    arg = sum / len(gd_degree_centrality)
    degrees.append(arg)
        
# 固有ベクトル集中度
def eigenvector_concentrate(G):
    # 固有ベクトル中心性
    gd_eigenvector_centrality = pd.Series(nx.eigenvector_centrality(G))
    sum = 0
    for i in range(0, len(gd_eigenvector_centrality)):
        sum += gd_eigenvector_centrality[i]
    arg = sum / len(gd_eigenvector_centrality)
    eigenvectors.append(arg)
        
# ページランク集中度
def pagerank_concentrate(G):
    # ページランク：PageRank
    gd_pagerank = pd.Series(nx.pagerank(G))
    sum = 0
    for i in range(0, len(gd_pagerank)):
        sum += gd_pagerank[i]
    arg = sum / len(gd_pagerank)
    pageranks.append(arg)
        
# 媒介集中度
def betweeness_concentrate(G):
    # 媒介中心性：Betweeness centrality
    gd_betweenness_centrality = pd.Series(nx.betweenness_centrality(G))
    sum = 0
    for i in range(0, len(gd_betweenness_centrality)):
        sum += gd_betweenness_centrality[i]
    arg = sum / len(gd_betweenness_centrality)
    betweenesses.append(arg)

def make_network(n, p, q):

    nodes = []
    edges = []
    clusterings = []
    lengths = []
    
    for i in range(0, q):

        # G
        G = nx.erdos_renyi_graph(n, p)

        # V
        nodes.append(G.number_of_nodes())

        # E
        edges.append(G.number_of_edges())

        # C
        clusterings.append(nx.average_clustering(G))

        # D
        lengths.append(nx.average_shortest_path_length(G))

        # 次数集中度
        degree_concentrate(G)

        # 固有ベクトル集中度
        eigenvector_concentrate(G)

        # ページランク集中度
        pagerank_concentrate(G)

        # 媒介集中度
        betweeness_concentrate(G) 

    # excelファイルに書き込む 
    # 以下、エクセル出力に関する部分
    import openpyxl
    
    wb = openpyxl.Workbook() #エクセルファイルを新規作成
    sheet = wb.active
    sheet.title = 'dataset' 
    
    sheet["A1"].value = '頂点数'
    sheet["B1"].value = '辺の数'
    sheet["C1"].value = 'クラスター係数'
    sheet["D1"].value = '平均距離'
    sheet["E1"].value = '次数集中度'
    sheet["F1"].value = '固有ベクトル集中度'
    sheet["G1"].value = 'ページランク集中度'
    sheet["H1"].value = '媒介集中度'
    
    for i in range(1, q+1):
        sheet.cell(column=1, row=i+1, value=nodes[i-1])
        sheet.cell(column=2, row=i+1, value=edges[i-1])
        sheet.cell(column=3, row=i+1, value=clusterings[i-1])
        sheet.cell(column=4, row=i+1, value=lengths[i-1])
        sheet.cell(column=5, row=i+1, value=degrees[i-1])
        sheet.cell(column=6, row=i+1, value=eigenvectors[i-1])
        sheet.cell(column=7, row=i+1, value=pageranks[i-1])
        sheet.cell(column=8, row=i+1, value=betweenesses[i-1])

    wb.save('ER_dataset_excel.xlsx')
    wb.close()

v = int(input("V:"))
p = float(input("P:"))
count = int(input("Count:"))
make_network(v,p,count)