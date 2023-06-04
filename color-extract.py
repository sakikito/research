import matplotlib.pyplot as plt #画像を表示するためのモジュール
import cv2  #OpenCVをインポート
import numpy as np #numpyをインポート

img = cv2.imread("ramen.jpeg") #画像の読み込み
img_array = np.asarray(img) #numpyで扱える配列をつくる

R = []
G = []
B = []

for x in range(0, len(img_array)):
    for y in range(0, len(img_array)):
        R.append(img_array[x][y][0])
        G.append(img_array[x][y][1])
        B.append(img_array[x][y][2])
 
# excelファイルに書き込む

tmp_num = len(img_array)
 
##以下、エクセル出力に関する部分
import openpyxl
 
wb = openpyxl.Workbook() #エクセルファイルを新規作成
sheet = wb.active
sheet.title = 'ramen' 

sheet["A1"].value = 'ラーメン'
sheet["A2"].value = 'R'
sheet["B2"].value = 'G'
sheet["C2"].value = 'B'
 
for i in range(2, tmp_num+2):
    sheet.cell(column=1, row=i+1, value=R[i-1])
    sheet.cell(column=2, row=i+1, value=G[i-1])
    sheet.cell(column=3, row=i+1, value=B[i-1])

wb.save('ramen_RGB_excel.xlsx')
wb.close()