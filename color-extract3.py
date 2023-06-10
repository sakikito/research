import matplotlib.pyplot as plt #画像を表示するためのモジュール
import cv2  #OpenCVをインポート
import numpy as np #numpyをインポート

# image
# 画像数 + 1
img_num = 5

for i in range(1, img_num):

    img = cv2.imread("resize{}.jpg".format(i)) #画像の読み込み
    img_array = np.asarray(img) #numpyで扱える配列をつくる

    R = []
    G = []
    B = []

    for x in range(0, len(img_array)):
        for y in range(0, len(img_array)):
            R.append(img_array[x][y][0])
            G.append(img_array[x][y][1])
            B.append(img_array[x][y][2])
 
    # excelファイルに書き込む

    tmp_num = len(img_array)
    
    ##以下、エクセル出力に関する部分
    import openpyxl
    
    wb = openpyxl.Workbook() #エクセルファイルを新規作成
    sheet = wb.active
    sheet.title = 'mentaiko{}'.format(i) 

    sheet["A1"].value = 'mentaiko or tarako'.format(i)
    sheet["A2"].value = 'R'
    sheet["B2"].value = 'G'
    sheet["C2"].value = 'B'
    
    for j in range(2, tmp_num+2):
        sheet.cell(column=1, row=j+1, value=R[j-1])
        sheet.cell(column=2, row=j+1, value=G[j-1])
        sheet.cell(column=3, row=j+1, value=B[j-1])

    wb.save('mentaiko{}_RGB_excel.xlsx'.format(i))
    wb.close()